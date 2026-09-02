from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell

from app.data.manifest import DatasetEntry


class ExcelValidationError(ValueError):
    pass


SCHEMA_NAME_COLUMN = 1  # '컬럼명' 열 index in schema sheet (0-based: column B)


def _json_safe_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str | bool | int | float):
        return value
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def classify_cell_value(cell: Cell | openpyxl.cell.cell.MergedCell) -> tuple[Any, str | None]:
    """셀 값과 value_state 분류 — 0/빈값/문자열0 구분"""
    value = cell.value
    if value is None:
        return None, "empty"
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None, "empty"
        if stripped == "0":
            return stripped, "string_zero"
        return value, None
    if isinstance(value, bool):
        return value, None
    if isinstance(value, int):
        if value == 0:
            return 0, "numeric_zero"
        return value, None
    if isinstance(value, float):
        if value == 0.0:
            return 0.0, "numeric_zero"
        return value, None
    if isinstance(value, Decimal):
        if value == 0:
            return _json_safe_value(value), "numeric_zero"
        return _json_safe_value(value), None
    return _json_safe_value(value), None


def read_schema_columns(path: Path, sheet_name: str) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            msg = f"schema sheet not found: {sheet_name}"
            raise ExcelValidationError(msg)
        worksheet = workbook[sheet_name]
        columns: list[str] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            column_name = row[SCHEMA_NAME_COLUMN]
            if column_name is None:
                continue
            columns.append(str(column_name))
        if not columns:
            msg = "schema sheet has no column definitions"
            raise ExcelValidationError(msg)
        return columns
    finally:
        workbook.close()


def read_data_headers(path: Path, sheet_name: str) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            msg = f"data sheet not found: {sheet_name}"
            raise ExcelValidationError(msg)
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(value) for value in header_row if value is not None]
        if not headers:
            msg = "data sheet header row is empty"
            raise ExcelValidationError(msg)
        return headers
    finally:
        workbook.close()


@dataclass(frozen=True)
class ParsedDataRow:
    row_number: int
    payload: dict[str, Any]
    value_states: dict[str, str]


def iter_data_rows(
    path: Path,
    entry: DatasetEntry,
    headers: list[str],
) -> Iterator[ParsedDataRow]:
    """openpyxl read_only=False로 셀 타입·선행0 보존"""
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        worksheet = workbook[entry.data_sheet]
        max_row = worksheet.max_row or 1
        for row_idx in range(2, max_row + 1):
            cells = [
                worksheet.cell(row=row_idx, column=col_idx + 1)
                for col_idx in range(len(headers))
            ]
            if all(cell.value is None for cell in cells):
                continue
            payload: dict[str, Any] = {}
            value_states: dict[str, str] = {}
            for header, cell in zip(headers, cells, strict=True):
                safe_value, state = classify_cell_value(cell)
                payload[header] = _json_safe_value(safe_value)
                if state is not None:
                    value_states[header] = state
            yield ParsedDataRow(row_number=row_idx, payload=payload, value_states=value_states)
    finally:
        workbook.close()


def count_data_rows(path: Path, entry: DatasetEntry) -> int:
    return sum(1 for _ in iter_data_rows(path, entry, read_data_headers(path, entry.data_sheet)))
