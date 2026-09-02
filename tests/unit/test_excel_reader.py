from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from app.data.excel_reader import classify_cell_value, iter_data_rows, read_data_headers
from app.data.manifest import DatasetEntry


@pytest.fixture
def mini_dataset(tmp_path: Path) -> tuple[Path, DatasetEntry]:
    schema_path = tmp_path / "mini_schema.xlsx"
    data_path = tmp_path / "mini_data.xlsx"

    schema_wb = openpyxl.Workbook()
    schema_ws = schema_wb.active
    schema_ws.title = "schema"
    schema_ws.append(["순번", "컬럼명", "데이터타입", "Nullable", "컬럼코멘트"])
    for idx, col in enumerate(["pd_no", "amount", "code", "mat_dt"], start=1):
        schema_ws.append([str(idx), col, "text", "YES", col])
    schema_wb.save(schema_path)

    data_wb = openpyxl.Workbook()
    data_ws = data_wb.active
    data_ws.title = "data"
    data_ws.append(["pd_no", "amount", "code", "mat_dt"])
    data_ws.append(["B001", 0, "0", 99991231])
    data_ws.append(["B002", None, "", None])
    data_wb.save(data_path)

    entry = DatasetEntry(
        logical_table="TEST",
        raw_table="raw_bond_kr",
        schema_file="mini_schema.xlsx",
        data_file="mini_data.xlsx",
        schema_sha256="0" * 64,
        data_sha256="0" * 64,
        schema_sheet="schema",
        data_sheet="data",
        expected_rows=2,
        expected_columns=4,
        source_key_fields=["pd_no"],
    )
    return tmp_path, entry


def test_classify_cell_value_states() -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet["A1"] = None
    worksheet["A2"] = 0
    worksheet["A3"] = "0"

    assert classify_cell_value(worksheet["A1"]) == (None, "empty")
    assert classify_cell_value(worksheet["A2"])[1] == "numeric_zero"
    assert classify_cell_value(worksheet["A3"]) == ("0", "string_zero")
    workbook.close()


def test_iter_data_rows_preserves_zero_and_empty(mini_dataset: tuple[Path, DatasetEntry]) -> None:
    base, entry = mini_dataset
    headers = read_data_headers(base / entry.data_file, entry.data_sheet)
    rows = list(iter_data_rows(base / entry.data_file, entry, headers))
    assert len(rows) == 2

    first = rows[0].payload
    assert first["amount"] == 0
    assert first["code"] == "0"
    assert first["mat_dt"] == 99991231
    assert rows[0].value_states["amount"] == "numeric_zero"
    assert rows[0].value_states["code"] == "string_zero"

    second = rows[1].payload
    assert second["amount"] is None
    assert second["code"] is None
    assert "amount" in rows[1].value_states
